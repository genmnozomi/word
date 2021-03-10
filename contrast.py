import MySQLdb
from openpyxl import load_workbook

list1 = []  # 待数据输入
table = 'test'  # 数据表

def db_con():
    """
    数据库连接
    """
    db = MySQLdb.connect('localhost', 'root', 'abc123,.', 'japanese', charset='utf8')
    return db

def table_query(table):
    """
    数据表查询
    """
    db = db_con()
    cursor = db.cursor()
    sql = f'select * from {table}'
    cursor.execute(sql)
    data = cursor.fetchall()
    db.close()
    return data

def add_list():
    """
    把获取的数据循环添加字典dict1后加入元组list1里面
    """
    data = table_query(table)
    global list1
    for i in data:
        dict1 = dict()
        dict1['id'] = i[0]
        dict1['word'] = i[1]
        dict1['alias'] = i[2]
        dict1['chinese'] = i[3]
        dict1['tone'] = i[4]
        dict1['wtype'] = i[5]
        list1.append(dict1)
    #print(list1)

#add_list()
#print(list1)

def con_trast():
    """
    从数据表获取出来的数据进行对比，这里显示word数据，输入alias对比是否相等
    """
    add_list()
    i = 0
    while i < len(list1):
        print(list1[i]['word'])
        in_alias = input('请输入单词假名')
        if in_alias == list1[i]['alias']:
            i += 1
        else:
            print('alias error')

#con_trast()

inbook = 'word.xlsx'    # excel文件
insheet = 'Sheet1'      # 选择表格


def xlsx_insert():
    """
    获取excel表里面的数据添加到数据表里面
    """
    book = load_workbook(inbook)
    sheet = book[insheet]
    data = sheet.iter_rows()

    db = db_con()
    cursor = db.cursor()
    sql = 'insert into test (word, alias, chinese, tone, wtype) values (%s, %s, %s, %s, %s)'

    for i in data:
        word = i[0].value
        alias = i[1].value
        chinese = i[2].value
        tone = i[3].value
        wtype = i[4].value
        values = (word, alias, chinese, tone, wtype)
        cursor.execute(sql, values)

    cursor.close()
    db.commit()
    db.close()

xlsx_insert()
add_list()
print(list1)